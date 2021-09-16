'''textFiletoExcel.py reads contents of several text files and insert those contents into spreadsheet with one line of text per row.'''

import openpyxl
from openpyxl.styles import Font

def textFileToExcel(*filenames):
    wb = openpyxl.Workbook()
    ws = wb.active
    bold = Font(bold=True)
    col = 1
    try:
        for filename in filenames:
            textfile = open(filename, 'r')
            listOflines = textfile.readlines()
            ws.cell(row=1, column=col).value = textfile.name
            ws.cell(row=1, column=col).font = bold
            if listOflines == []:
                ws.cell(row=2, column=col).value = 'This file was empty'
                ws.cell(row=2, column=col).font = bold
            else:
                for i in range(len(listOflines)):
                    ws.cell(row=i+2, column=col).value = listOflines[i]
            col += 1
    except FileNotFoundError:
        print('File not found. Check names of given filenames and try again.')
    wb.save('textFileToExcel.xlsx')    



