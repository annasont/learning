'''updateProduceSales.py updates prices of products in produceSales.xlsx'''
import openpyxl
from openpyxl.styles import Font
wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

'''my way'''
# for row in range(2, sheet.max_row + 1):
#     product = sheet['A' + str(row)].value
#     if product == 'Garlic':
#         sheet['B' + str(row)] = 2.15
#     if product == 'Lemon':
#         sheet['B' + str(row)] = 4.30
# wb.save('produceSalesCopy.xlsx')

'''better way'''
updatePrices = {'Garlic': 1.30, 'Lemon': 2.07}
bold = Font(bold=True)

for row in range (2, sheet.max_row + 1):
    product = sheet['A' + str(row)].value
    if product in updatePrices:
        sheet['B' + str(row)] = updatePrices[product]
        sheet['B' + str(row)].font = bold

wb.save('produceSalesCopy.xlsx')





