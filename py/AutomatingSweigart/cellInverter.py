'''cellInverter.py inverts the row and column of the cells in the spreadsheet.'''

import openpyxl

def invertCells(filename):

    #loading workbook
    wb = openpyxl.load_workbook(filename)
    ws = wb.active

    #saving data from workbook to list
    wsContent = []
    for r in range(ws.max_row):
        wsContent.append([])
        for c in range(ws.max_column):
            wsContent[r].append(ws.cell(row=r+1, column=c+1).value)

    #opening new workbook
    newWb = openpyxl.Workbook()
    newWs = newWb.active

    #inverting rows with columns
    for c in range(ws.max_column):
        for r in range(ws.max_row):
            newWs.cell(row=c+1, column=r+1).value = wsContent[r][c]
    #saving in new file
    newWb.save('inverted.xlsx')

invertCells('pp.xlsx')