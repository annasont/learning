'''excelToCvs.py converts all excel files in current woriking directory to csv format '''
import os, openpyxl, csv
from openpyxl.utils.cell import get_column_letter

#creating new folder
try:
    os.mkdir(os.path.join('.', 'csvFiles'))
    print('Created folder "csvFiles".')
except:
    print('Could not create new folder named "csvFiles".')

#only files with .xlsx extension:
allExcelFiles = []
for filename in os.listdir():
    if filename.endswith('.xlsx'):
        allExcelFiles.append(filename)
print('Searching current directory...')

try:
    #going through all excel files, all sheets
    for filename in allExcelFiles:
        wb = openpyxl.load_workbook(filename)
        allSheetNames = wb.sheetnames
        for sheet in allSheetNames:
            #creating new file for every sheet
            outputFile = open(os.path.join('.', 'csvFiles', '%s-%s.csv' % (filename[:-5], sheet)), 'w', newline='')
            outputWriter = csv.writer(outputFile)
            ws = wb[sheet]
            #adding values from row to list
            rowValues = []
            for row in (ws['A1': get_column_letter(ws.max_column) + str(ws.max_row)]):
                for cell in row:
                    rowValues.append(cell.value)
                #adding row to csv file
                outputWriter.writerow(rowValues)
            outputFile.close()
    print('Converted all files to cvs.')
except:
    print('Something went wrong. Could not make copy.')
   
            
