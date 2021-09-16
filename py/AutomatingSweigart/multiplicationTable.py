'''multiplicationTable.py takes number n from command line and creates an n x n multiplication table.'''

import sys, openpyxl
from openpyxl.styles import Font
from openpyxl.utils.cell import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Multiplication table'

#takes number from command line
number = sys.argv[1]

#creating labels
n = 1
bold = Font(bold=True)
for i in range(2, int(number) + 2):
    sheet['A' + str(i)].value = n
    sheet['A' + str(i)].font = bold
    sheet.cell(row=1, column=i).value = n
    sheet.cell(row=1, column=i).font = bold
    n += 1

#filling out table with formulas
lastCell = get_column_letter(int(number)+1) + str(int(number) + 1)
aColumn = 2 #starts with A2
for row in (sheet['B2': lastCell]):
    bRow = 2 #starts with B1
    for c in row: 
        sheet[c.coordinate].value = '=%s%s*%s%s' % (get_column_letter(bRow), '1','A', aColumn)
        bRow += 1
    aColumn += 1

sheet.freeze_panes = 'B2'

wb.save('MultiplicationTable.xlsx')

