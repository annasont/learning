'''sendReminders.py sends emails based on payment status in spreadsheet.'''
import openpyxl, smtplib
from openpyxl.utils.cell import get_column_letter

def sendReminders(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # creating dictionary:
    # memberData[name]:
    # memberData[name][email] = ''
    # memberData[mame][due] = []
    memberData = {}
    for row in range(2, sheet.max_row + 1):
        name = sheet['A' + str(row)].value
        memberData.setdefault(name, {})
        memberData[name]['email'] = sheet['B' + str(row)].value

        dueMonths = []
        for i in range(3, sheet.max_column + 1):
            if sheet[get_column_letter(i) + str(row)].value == None:
                dueMonths.append(sheet[get_column_letter(i) + '1'].value)
        memberData[name]['due'] = dueMonths

    # Sending email to members with 'due'
    smtpObj = smtplib.SMTP('smtp.gmail.com', 587)
    smtpObj.ehlo()
    smtpObj.starttls()
    print('Enter password:')
    password = input()
    smtpObj.login('aniasontowska@gmail.com', password)

    membersWithDue = []
    for member in memberData:
        if memberData[member]['due'] != []:
            missingPayment = ', '.join(memberData[member]['due'])
            message = 'Subject: Reminder about missing payments\nDear %s,\nWe would like to inform you that payment for following month(s) is missing: %s.\n\nBest regards' % (member, missingPayment)
            sendMailStatus = smtpObj.sendmail('aniasontowska@gmail.com', memberData[member]['email'], message)
            if sendMailStatus != {}:
                print('There was a problem sending email to %s: %s' % (memberData[member]['email'], sendMailStatus))
            else:
                membersWithDue.append('%s => %s' % (member, memberData[member]['email']))
    print('Email sendt to following resepients:\n%s' % '\n'.join(membersWithDue))

    smtpObj.quit()

print(sendReminders('duesRecords.xlsx'))