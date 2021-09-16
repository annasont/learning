'''fromExcelToTextFile.py'''
import openpyxl
from openpyxl.utils.cell import get_column_letter

def fromExcelToTextFile(ecxelFile):
    wb = openpyxl.load_workbook(ecxelFile)
    ws = wb.active

    for col in range(1, ws.max_column + 1):
        filename = ws.cell(row=1, column=col).value
        textfile = open(filename, 'w')
        for cell in ws[get_column_letter(col)][1:]:
            if not cell.value is None:
                textfile.write(cell.value)
        textfile.close()

fromExcelToTextFile('textFileToExcel.xlsx')